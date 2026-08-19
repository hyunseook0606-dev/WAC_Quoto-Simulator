# -*- coding: utf-8 -*-
"""Read Master_DB from WAC Excel simulator and calculate desk quotations."""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
DEFAULT_XLSX = ROOT / "WAC_Air_Quotation_Simulator.xlsx"


def _num(v: Any, fallback: float = 0.0) -> float:
    if v is None or v == "":
        return fallback
    try:
        n = float(v)
    except (TypeError, ValueError):
        return fallback
    return n if n == n else fallback  # NaN guard


@dataclass
class AirRate:
    route: str
    min_amt: float
    r_under45: float
    r45: float
    r100: float
    r500: float
    r1000: float
    fsc: float
    ssc: float
    currency: str = "HKD"


@dataclass
class LocalRate:
    item: str
    unit: str
    rate: float
    min_amt: float


@dataclass
class Master:
    vol_factor: float
    cbm_divisor: float
    wb45: float
    wb100: float
    wb500: float
    wb1000: float
    air: list[AirRate]
    local: list[LocalRate]
    source: str


@dataclass
class QuoteLine:
    label: str
    group: str
    amount: float
    note: str = ""


@dataclass
class QuoteResult:
    route: str
    cbm: float
    cw: float
    break_label: str
    air_rate: float
    currency: str
    fx: float
    lines: list[QuoteLine] = field(default_factory=list)
    total: float = 0.0

    def to_dict(self) -> dict[str, Any]:
        return {
            "route": self.route,
            "cbm": round(self.cbm, 4),
            "cw": round(self.cw, 2),
            "break_label": self.break_label,
            "air_rate_per_kg": round(self.air_rate, 2),
            "currency": self.currency,
            "fx": self.fx,
            "lines": [
                {
                    "label": ln.label,
                    "group": ln.group,
                    "amount": round(ln.amount, 2),
                    "note": ln.note,
                }
                for ln in self.lines
            ],
            "total_approx": round(self.total, 2),
        }


def break_label(cw: float, master: Master) -> tuple[str, float]:
    if cw < master.wb45:
        return "-45", 0.0  # rate filled by caller
    if cw >= master.wb1000:
        return "+1000", 0.0
    if cw >= master.wb500:
        return "+500", 0.0
    if cw >= master.wb100:
        return "+100", 0.0
    return "+45", 0.0


def pick_air_rate(air: AirRate, label: str) -> float:
    return {
        "-45": air.r_under45,
        "+45": air.r45,
        "+100": air.r100,
        "+500": air.r500,
        "+1000": air.r1000,
    }[label]


def pick_local(master: Master, name: str) -> LocalRate | None:
    for row in master.local:
        if row.item == name:
            return row
    return None


def load_master(xlsx_path: Path | str = DEFAULT_XLSX) -> Master:
    path = Path(xlsx_path)
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb["Master_DB"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    def cell(r: int, c: int) -> Any:
        # r,c are 0-based
        if r < 0 or r >= len(rows):
            return None
        row = rows[r]
        if c < 0 or c >= len(row):
            return None
        return row[c]

    air: list[AirRate] = []
    air_header = -1
    for r in range(len(rows)):
        if str(cell(r, 0) or "").strip().upper() == "ROUTE":
            air_header = r
            break

    if air_header >= 0:
        for r in range(air_header + 1, len(rows)):
            route = str(cell(r, 0) or "").strip()
            if not route or route.startswith("2.") or "Local" in route:
                break
            if "-" not in route:
                break
            cur = str(cell(r, 9) or "HKD").strip().upper()
            if cur not in ("USD", "HKD"):
                origin = route.split("-")[0].upper()
                cur = "HKD" if origin == "HKG" else "USD"
            air.append(
                AirRate(
                    route=route,
                    min_amt=_num(cell(r, 1)),
                    r_under45=_num(cell(r, 2)),
                    r45=_num(cell(r, 3)),
                    r100=_num(cell(r, 4)),
                    r500=_num(cell(r, 5)),
                    r1000=_num(cell(r, 6)),
                    fsc=_num(cell(r, 7)),
                    ssc=_num(cell(r, 8)),
                    currency=cur,
                )
            )

    local: list[LocalRate] = []
    local_header = -1
    for r in range(len(rows)):
        if str(cell(r, 0) or "").strip() == "Charge Item":
            local_header = r
            break

    if local_header >= 0:
        for r in range(local_header + 1, len(rows)):
            item = str(cell(r, 0) or "").strip()
            if not item or item.startswith("구간") or item.startswith("행"):
                break
            local.append(
                LocalRate(
                    item=item,
                    unit=str(cell(r, 1) or ""),
                    rate=_num(cell(r, 2)),
                    min_amt=_num(cell(r, 3)),
                )
            )

    if not air:
        raise ValueError(f"No air routes in Master_DB ({path.name})")

    return Master(
        vol_factor=_num(cell(4, 1), 167),
        cbm_divisor=_num(cell(5, 1), 1_000_000),
        wb45=_num(cell(8, 1), 45),
        wb100=_num(cell(9, 1), 100),
        wb500=_num(cell(10, 1), 500),
        wb1000=_num(cell(11, 1), 1000),
        air=air,
        local=local,
        source=path.name,
    )


def calc_piece_cw(
    length: float,
    width: float,
    height: float,
    qty: int,
    gross: float,
    master: Master,
) -> tuple[float, float]:
    cbm = length * width * height * qty / master.cbm_divisor
    cw = max(gross, cbm * master.vol_factor)
    return cbm, cw


def calc_quote(
    *,
    route: str,
    length: float,
    width: float,
    height: float,
    qty: int = 1,
    gross: float,
    bl_count: int = 1,
    fx: float = 1.0,
    exceptions: dict[str, float | None] | None = None,
    master: Master | None = None,
    xlsx_path: Path | str = DEFAULT_XLSX,
) -> QuoteResult:
    """Desk-style quote: Master auto lines + per-job exceptions (Excel J-column)."""
    master = master or load_master(xlsx_path)
    route = route.strip().upper()
    air_row = next((a for a in master.air if a.route == route), None)
    if not air_row:
        raise ValueError(f"No Master rate for route {route}")

    exc = {k.lower(): v for k, v in (exceptions or {}).items()}
    cbm, cw = calc_piece_cw(length, width, height, qty, gross, master)
    br, _ = break_label(cw, master)
    air_rate = pick_air_rate(air_row, br)

    def use(key: str, computed: float) -> float:
        if key in exc and exc[key] is not None:
            return float(exc[key])
        return computed

    air_amt = max(air_rate * cw, air_row.min_amt)
    fsc = air_row.fsc * cw
    ssc = air_row.ssc * cw

    handling_m = pick_local(master, "Handling Fee")
    doc_m = pick_local(master, "Doc Fee")
    truck_m = pick_local(master, "Trucking")
    term_m = pick_local(master, "Terminal Charge")
    cfs_m = pick_local(master, "CFS")
    pickup_m = pick_local(master, "Pickup (temp)")
    export_m = pick_local(master, "Export declaration")
    repack_m = pick_local(master, "RE-PACKING")
    xray_m = pick_local(master, "XRAY")
    gate_m = pick_local(master, "Gate / parking / toll")

    handling = use("handling", max(handling_m.rate, handling_m.min_amt) if handling_m else 0)
    doc = use("doc", (max(doc_m.rate, doc_m.min_amt) * bl_count) if doc_m else 0)
    trucking = use(
        "trucking",
        max(truck_m.rate * cbm, truck_m.min_amt) if truck_m else 0,
    )
    terminal = use(
        "terminal",
        max(term_m.rate * cw, term_m.min_amt) if term_m else 0,
    )
    xray = use("xray", (xray_m.rate * cw) if xray_m else 0)
    cfs = use("cfs", max(cfs_m.rate * cw, cfs_m.min_amt) if cfs_m else 0)
    pickup = use("pickup", pickup_m.min_amt if pickup_m else 0)
    export = use(
        "export",
        (max(export_m.rate * bl_count, export_m.min_amt) * bl_count)
        if export_m
        else 0,
    )
    repack = use("repack", (repack_m.rate * qty) if repack_m else 0)
    gate = use("gate", gate_m.min_amt if gate_m else 0)
    palletizing = use("palletizing", float(exc.get("palletizing", 0) or 0))
    other = use("other", float(exc.get("other", 0) or 0))

    lines = [
        QuoteLine("Air Freight", "air", air_amt, f"{br} @ {air_rate}/kg"),
        QuoteLine("FSC", "air", fsc),
        QuoteLine("SSC", "air", ssc),
        QuoteLine("Handling Fee", "local", handling),
        QuoteLine("Doc Fee", "local", doc),
        QuoteLine("Trucking", "local", trucking),
        QuoteLine("Terminal Charge", "local", terminal),
        QuoteLine("XRAY", "local", xray),
        QuoteLine("CFS", "local", cfs),
        QuoteLine("Pickup (temp)", "variable", pickup),
        QuoteLine("Export declaration", "local", export),
        QuoteLine("RE-PACKING", "variable", repack),
        QuoteLine("Gate / parking / toll", "variable", gate),
    ]
    if palletizing > 0:
        lines.append(QuoteLine("Palletizing", "variable", palletizing))
    if other > 0:
        lines.append(QuoteLine("Other", "variable", other))

    subtotal = sum(ln.amount for ln in lines)
    total = subtotal * (fx if air_row.currency == "HKD" else 1.0)

    return QuoteResult(
        route=route,
        cbm=cbm,
        cw=cw,
        break_label=br,
        air_rate=air_rate,
        currency=air_row.currency,
        fx=fx,
        lines=lines,
        total=total,
    )
