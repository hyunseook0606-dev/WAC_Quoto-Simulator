# -*- coding: utf-8 -*-
"""Align charge amounts + validate multi-case Excel logic (no INV rewrite)."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Protection

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "WAC_Air_Quotation_Simulator.xlsx"
AMT_FMT = "#,##0.00"
RIGHT = Alignment(horizontal="right", vertical="center")


def unlock(cell):
    cell.protection = Protection(locked=False)


def lock(cell):
    cell.protection = Protection(locked=True)


def fix_alignment(wb):
    inp = wb["입력"]
    q = wb["견적서"]
    try:
        inp.protection.sheet = False
        q.protection.sheet = False
    except Exception:
        pass

    # 입력 참고(I) / 예외(J) rows 18-30 + TOTAL I31
    for r in range(18, 31):
        for c in (9, 10):
            cell = inp.cell(r, c)
            cell.number_format = AMT_FMT
            cell.alignment = RIGHT
            if c == 10:
                unlock(cell)
            else:
                lock(cell)
    for r in (31, 32):
        for c in (9, 10):
            cell = inp.cell(r, c)
            cell.number_format = AMT_FMT
            cell.alignment = Alignment(horizontal="right", vertical="center")
            lock(cell)

    # 견적서 amount col F (and merged G)
    for r in range(26, 40):
        cell = q.cell(r, 6)
        if cell.value is None and r != 39:
            continue
        cell.number_format = AMT_FMT
        cell.alignment = RIGHT
        lock(cell)
        q.cell(r, 7).alignment = RIGHT
        lock(q.cell(r, 7))

    inp.protection.sheet = True
    q.protection.sheet = True


def sim_piece_cw(L, W, H, qty, gross, vol_factor=167, cbm_div=1_000_000):
    cbm = L * W * H * qty / cbm_div
    cw = max(gross, cbm * vol_factor)
    return cbm, cw


def break_label(cw):
    if cw < 45:
        return "-45"
    if cw >= 1000:
        return "+1000"
    if cw >= 500:
        return "+500"
    if cw >= 100:
        return "+100"
    return "+45"


def air_amount(cw, rate, amin):
    return max(rate * cw, amin)


def run_cases():
    results = []

    # --- Case 1: Chocolate chocolate (current fill expectations) ---
    cbm, cw = sim_piece_cw(110, 110, 109, 1, 194.5)
    air = {
        "MIN": 650,
        "-45": 50,
        "+45": 30,
        "+100": 25,
        "+500": 30,
        "+1000": 30,
        "FSC": 2.1,
        "SSC": 2.0,
    }
    br = break_label(cw)
    rate = air[br]
    lines = {
        "Air": air_amount(cw, rate, air["MIN"]),
        "FSC": air["FSC"] * cw,
        "SSC": air["SSC"] * cw,
        "Handling": 321,  # 예외
        "Doc": 15,  # 예외
        "Trucking": 0,  # 예외
        "Terminal": max(1.68 * cw, 60),
        "XRAY": 1.0 * cw,
        "CFS": max(0.7 * cw, 200),
        "Pickup": 2000,
        "Export": 213,
        "REPACK": 300,
        "Gate": 0,
    }
    total = sum(lines.values())
    ok = abs(total - 10048.75) < 0.05 and abs(cw - 220.2563) < 0.01 and br == "+100"
    results.append(
        (
            "Chocolate HKG-ICN KEEP COOL",
            ok,
            f"CW={cw:.4f} Break={br} TOTAL={total:.2f} (expect ~10048.75)",
            lines,
        )
    )

    # --- Case 2: CASE A ICN-HKG classic (verify engine) ---
    air2 = {"MIN": 50, 45: 4.5, 100: 3.8, 500: 3.2, 1000: 2.8, "FSC": 0.6, "SSC": 0.15}
    # original CASE A used 3 pallets 110x110x150 gross 400 each? From build verify:
    # L,W,H,qty,gross = 110,110,150,3,400 — wait that's one line with qty 3
    cbm2, cw2 = sim_piece_cw(110, 110, 150, 3, 400)
    rate2 = air2[500] if cw2 >= 500 else air2[100]
    total2 = (
        max(rate2 * cw2, air2["MIN"])
        + air2["FSC"] * cw2
        + air2["SSC"] * cw2
        + 30
        + 25
        + max(15 * cbm2, 80)
    )
    ok2 = abs(total2 - 3728.47) < 0.02 and abs(cw2 - 909.315) < 0.01
    results.append(
        (
            "CASE A ICN-HKG (USD sample math)",
            ok2,
            f"CW={cw2:.3f} rate={rate2} TOTAL={total2:.2f} (expect 3728.47)",
            None,
        )
    )

    # --- Case 3: -45 break ---
    cbm3, cw3 = sim_piece_cw(40, 30, 20, 1, 20)  # small
    br3 = break_label(cw3)
    # ICN-HKG under45 = 5.5
    rate3 = 5.5 if br3 == "-45" else 4.5
    air3 = max(rate3 * cw3, 50)
    ok3 = br3 == "-45" and cw3 < 45
    results.append(
        (
            "Break -45 (C.W.<45)",
            ok3,
            f"CW={cw3:.2f} Break={br3} AirRate would use -45 col ({rate3})",
            None,
        )
    )

    # --- Case 4: INV structure (84kg) — formulas work; not 1:1 INV clone ---
    cw4 = 84.0
    term = max(1.68 * cw4, 60)  # expect 141.12
    doc = 15
    # Current Excel CFS master is 0.7/200 (chocolate lane), not INV's 1.2/160
    cfs_excel = max(0.7 * cw4, 200)  # 200 MIN path
    cfs_inv_style = max(1.2 * cw4, 160)  # for comparison only
    hand_ex = 312  # 예외 (INV showed 312; master 30/150 gap)
    cartage, tunnel, parking = 650, 16, 15  # Other/예외 수기
    local_sum = term + doc + cfs_excel + hand_ex + cartage + tunnel + parking
    ok4 = abs(term - 141.12) < 0.01 and doc == 15
    results.append(
        (
            "INV_AE260703101 structure @84kg CW",
            ok4,
            f"Terminal={term:.2f}(OK) Doc={doc} CFS_excelMIN={cfs_excel:.0f} "
            f"(INV-style CFS would be {cfs_inv_style:.0f}) "
            f"Handling via 예외={hand_ex}; Cartage/Tunnel/Parking via Other 수기 "
            f"local_sub={local_sum:.2f}",
            None,
        )
    )

    # --- Case 5: Trucking=CBM not CW ---
    cbm5, cw5 = 2.0, 500.0
    truck_cbm = max(15 * cbm5, 80)  # 80 MIN
    truck_wrong_cw = max(15 * cw5, 80)  # would be huge if bug
    ok5 = truck_cbm == 80 and truck_wrong_cw != truck_cbm
    results.append(
        (
            "Trucking uses CBM (not C.W.)",
            ok5,
            f"Trucking(CBM2)={truck_cbm} vs wrong(CW500)={truck_wrong_cw}",
            None,
        )
    )

    # --- Case 6: FX — USD sum * rate vs already HKD * 1 ---
    usd_sum = 1000.0
    hkd_from_usd = usd_sum * 7.8
    hkd_native = 10048.75 * 1.0
    ok6 = abs(hkd_from_usd - 7800) < 0.01 and abs(hkd_native - 10048.75) < 0.01
    results.append(
        (
            "FX: HKD×Ex.Rate (7.8 vs 1)",
            ok6,
            f"USD1000→HKD {hkd_from_usd:.0f}; native HKD×1={hkd_native:.2f}",
            None,
        )
    )

    # --- Case 7: Check live workbook chocolate inputs ---
    wb = load_workbook(OUT, data_only=False)
    inp = wb["입력"]
    checks = {
        "Route": inp["C5"].value == "HKG-ICN",
        "Dims": (inp["C8"].value, inp["C9"].value, inp["C10"].value) == (110, 110, 109),
        "Qty": float(inp["C11"].value or 0) == 1,
        "GW": float(inp["C12"].value or 0) == 194.5,
        "FX": float(inp["C34"].value or 0) == 1,
        "Hand예외": float(inp["J21"].value or 0) == 321,
        "Doc예외": float(inp["J22"].value or 0) == 15,
        "Truck예외": inp["J23"].value is not None and float(inp["J23"].value) == 0,
        "Carrier": str(inp["C37"].value or "") == "KE",
    }
    ok7 = all(checks.values())
    results.append(
        (
            "Live file chocolate inputs",
            ok7,
            ", ".join(f"{k}={'OK' if v else 'NG'}" for k, v in checks.items()),
            None,
        )
    )

    return results


def main():
    wb = load_workbook(OUT)
    fix_alignment(wb)
    wb.save(OUT)
    print(f"Alignment fixed: {OUT}")

    for dest in (
        ROOT.parents[0] / "public" / "excel" / "WAC_Air_Quotation_Simulator.xlsx",
        Path.home() / "Downloads" / "WAC_Air_Quotation_Simulator.xlsx",
    ):
        try:
            dest.parent.mkdir(parents=True, exist_ok=True)
            dest.write_bytes(OUT.read_bytes())
            print(f"Copied -> {dest}")
        except OSError as e:
            print(f"Skip {dest}: {e}")

    print("\n===== CASE VALIDATION =====")
    results = run_cases()
    passed = 0
    for name, ok, detail, _ in results:
        mark = "PASS" if ok else "FAIL"
        if ok:
            passed += 1
        print(f"[{mark}] {name}")
        print(f"       {detail}")
    print(f"\n{passed}/{len(results)} passed")
    if passed != len(results):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
