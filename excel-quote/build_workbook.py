# -*- coding: utf-8 -*-
"""
WAC Air Quotation Simulator — template-patch builder

Project essence:
  Generic desk air-quotation tool (CM email UI).
  Master_DB = nearly-fixed rates.
  입력 = cargo + auto calc + 참고(Master) / 예외(manual).
  견적서 = 100% refs to 입력.
  Special cases (KEEP COOL / chocolate etc.) = fill 예외 + Other, NOT rewrite layout.

Approach:
  Copy clean baseline WAC_Air_Quotation_Simulator (2).xlsx
  then patch only Master / formulas / extra charge rows / FX / Guide.
"""

from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent
BASELINE = ROOT / "WAC_Air_Quotation_Simulator (2).xlsx"
OUT = ROOT / "WAC_Air_Quotation_Simulator.xlsx"
OUT_ALT = ROOT / "WAC_Air_Quotation_Simulator_v2.xlsx"

NAVY = "1A2A3A"
ORANGE = "F05023"
SLATE = "334155"
YELLOW = "FFF3D6"
GREEN_BG = "E8F5E9"
GREEN = "166534"
MUTED = "64748B"
LINE = "E2E8F0"
WHITE = "FFFFFF"
LIGHT = "F5F7F9"
INPUT_BG = "FFF7F2"

# Air table: A ROUTE … I SSC/kg | J CUR (USD/HKD)
AIR_RNG = "Master_DB!$A$16:$J$50"
CUR_COL = 10  # VLOOKUP col index for CUR
LOCAL_A = "Master_DB!$A$21:$A$40"
LOCAL_B = "Master_DB!$B$21:$B$40"
LOCAL_C = "Master_DB!$C$21:$C$40"
LOCAL_D = "Master_DB!$D$21:$D$40"

# Charge rows on 입력 after patch
# 18 Air, 19 FSC, 20 SSC, 21 Handling, 22 Doc, 23 Trucking
# 24 Terminal, 25-30 Other1-6, 31 TOTAL
CH_FIRST = 18
CH_LAST = 30
TOT_ROW = 31
TOT_AMT = "I31"
CUR_CELL = "C33"
FX_CELL = "C34"
# Quote meta (입력 footer → 견적서)
CARRIER_CELL = "C37"
REMARK_CELL = "C38"
VALID_CELL = "C39"
META_TIP_ROW = 41

thin = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)


def fill(c: str) -> PatternFill:
    return PatternFill("solid", fgColor=c)


def fnt(bold=False, size=11, color=NAVY):
    return Font(name="Malgun Gothic", bold=bold, size=size, color=color)


def unlock(cell):
    cell.protection = Protection(locked=False)


def lock(cell):
    cell.protection = Protection(locked=True)


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 10):
    for c in range(1, max_col + 1):
        s = ws.cell(src_row, c)
        d = ws.cell(dst_row, c)
        if s.has_style:
            d.font = copy(s.font)
            d.fill = copy(s.fill)
            d.border = copy(s.border)
            d.alignment = copy(s.alignment)
            d.number_format = s.number_format
            d.protection = copy(s.protection)


def unprotect(ws):
    try:
        ws.protection.sheet = False
    except Exception:
        pass


def protect(ws):
    ws.protection.sheet = True
    ws.protection.enable()


def patch_master(ws):
    """Insert -45 column, seed HKG-ICN, expand local items with short notes."""
    unprotect(ws)

    # Insert -45 between MIN and +45 (new column C)
    if ws.cell(15, 3).value != "-45":
        ws.insert_cols(3)
        ws.cell(15, 3, "-45")
        for c in range(1, 10):
            cell = ws.cell(15, c)
            cell.fill = fill(NAVY)
            cell.font = fnt(bold=True, size=10, color=WHITE)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin
            lock(cell)

        # Shifted seed values: old +45.. become cols 4..7, FSC/SSC 8/9
        # Fill -45 for existing routes (slightly above +45 as under-45 premium)
        for r in range(16, 28):
            route = ws.cell(r, 1).value
            if not route or not isinstance(route, str) or "-" not in route:
                continue
            plus45 = ws.cell(r, 4).value
            under = None
            if isinstance(plus45, (int, float)):
                under = round(float(plus45) + 1.0, 2)
            ws.cell(r, 3, under)
            for c in range(2, 10):
                cell = ws.cell(r, c)
                cell.fill = fill(YELLOW)
                cell.number_format = "0.00"
                cell.border = thin
                cell.alignment = Alignment(horizontal="center")
                unlock(cell)
            ws.cell(r, 1).fill = fill(INPUT_BG)
            ws.cell(r, 1).font = fnt(bold=True)
            unlock(ws.cell(r, 1))

    # Ensure HKG-ICN row exists (sample lane for special-case testing; not chocolate-locked)
    routes = []
    for r in range(16, 28):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and "-" in v:
            routes.append(v)
    if "HKG-ICN" not in routes:
        r = 16 + len(routes)
        # MIN, -45, +45, +100, +500, +1000, FSC, SSC  (case memo sample)
        vals = ("HKG-ICN", 650, 50, 30, 25, 30, 30, 2.1, 2.0)
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
            if c == 1:
                cell.font = fnt(bold=True)
                cell.fill = fill(INPUT_BG)
            else:
                cell.number_format = "0.00"
                cell.fill = fill(YELLOW)
            unlock(cell)
        routes.append("HKG-ICN")

    # Per-route currency (col J) — display only; rates stay in that currency's units
    _ensure_route_currency(ws)

    # Widen Note column slightly; keep notes SHORT
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["J"].width = 8

    # Local charges — continuous rows, no mid-table comment overlay
    # insert_cols(-45) shifts whole sheet: rewrite local header so Rate/MIN stay C/D
    local_hdr = ("Charge Item", "Unit", "Rate", "MIN", "Note")
    for c, v in enumerate(local_hdr, 1):
        cell = ws.cell(20, c, v)
        cell.fill = fill(NAVY)
        cell.font = fnt(bold=True, size=10, color=WHITE)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
        lock(cell)
    # Clear stray cells from column insert in header row
    for c in range(6, 10):
        ws.cell(20, c).value = None
        ws.cell(20, c).fill = PatternFill()
        ws.cell(20, c).border = Border()

    locals_ = [
        ("Handling Fee", "Per Shipment", 30, 30, "건별"),
        ("Doc Fee", "Per BL", 25, 25, "BL당"),
        ("Trucking", "Per CBM", 15, 80, "CBM, MIN"),
        ("Terminal Charge", "Per KG", 1.68, 60, "C.W. kg, MIN"),
        ("CFS", "Per KG", 0.70, 200, "C.W. kg, MIN"),
        ("Pickup (temp)", "Per Shipment", 0, 2000, "MIN/shpt"),
        ("Export declaration", "Per Entry", 200, 213, "MIN/entry"),
        ("RE-PACKING", "Per PLT", 300, 300, "Per PLT"),
        ("XRAY", "Per KG", 1.0, 0, "Per KG"),
        ("Gate / parking / toll", "Manual", 0, 0, "At cost"),
    ]
    # Clear old local area first (21-40)
    for r in range(21, 41):
        for c in range(1, 6):
            ws.cell(r, c).value = None
            ws.cell(r, c).fill = PatternFill()
            ws.cell(r, c).border = Border()

    for i, row in enumerate(locals_):
        r = 21 + i
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.border = thin
            cell.fill = fill(YELLOW)
            cell.alignment = Alignment(
                horizontal="left" if c in (1, 5) else "center",
                vertical="center",
            )
            unlock(cell)
            if c in (3, 4) and isinstance(v, (int, float)):
                cell.number_format = "0.00"
        ws.row_dimensions[r].height = 20

    tip_r = 21 + len(locals_) + 1
    # Unmerge tip area if a previous build left merges
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row == tip_r:
            try:
                ws.unmerge_cells(str(rng))
            except Exception:
                pass
    ws.merge_cells(start_row=tip_r, start_column=1, end_row=tip_r, end_column=8)
    tip_cell = ws.cell(
        tip_r,
        1,
        "A16~ ROUTE | J열 CUR=USD/HKD | -45 / +45… | Trucking=CBM · Terminal/CFS/XRAY=C.W.",
    )
    tip_cell.font = fnt(size=9, color=MUTED)
    tip_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[tip_r].height = 22
    lock(tip_cell)
    return routes


def _ensure_route_currency(ws):
    """Add Master col J = CUR (USD/HKD) per route. No insert_cols — append only."""
    hdr = ws.cell(15, CUR_COL, "CUR")
    hdr.fill = fill(NAVY)
    hdr.font = fnt(bold=True, size=10, color=WHITE)
    hdr.alignment = Alignment(horizontal="center", vertical="center")
    hdr.border = thin
    lock(hdr)

    # Default: ICN* → USD, HKG* → HKD (desk can edit yellow)
    for r in range(16, 28):
        route = ws.cell(r, 1).value
        if not isinstance(route, str) or "-" not in route:
            continue
        # Fill missing -45 if blank (legacy rows)
        if ws.cell(r, 3).value is None and isinstance(ws.cell(r, 4).value, (int, float)):
            ws.cell(r, 3, round(float(ws.cell(r, 4).value) + 1.0, 2))
            ws.cell(r, 3).fill = fill(YELLOW)
            ws.cell(r, 3).number_format = "0.00"
            ws.cell(r, 3).border = thin
            unlock(ws.cell(r, 3))

        cur = ws.cell(r, CUR_COL).value
        if cur not in ("USD", "HKD"):
            origin = route.split("-")[0].upper()
            cur = "HKD" if origin == "HKG" else "USD"
        cell = ws.cell(r, CUR_COL, cur)
        cell.fill = fill(YELLOW)
        cell.font = fnt(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
        unlock(cell)


def _footer_meta_row(ws, row, label, value, wide=False):
    """One yellow input row in footer: label B, value C:D (or C:J if wide)."""
    lab = ws.cell(row, 2, label)
    lab.font = fnt(bold=True, size=10)
    lab.fill = fill(LIGHT)
    lab.border = thin
    lab.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lock(lab)

    for rng in list(ws.merged_cells.ranges):
        if rng.min_row == row and rng.min_col >= 3:
            try:
                ws.unmerge_cells(str(rng))
            except Exception:
                pass

    end_col = 10 if wide else 4
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=end_col)
    cell = ws.cell(row, 3, value)
    cell.fill = fill(YELLOW)
    cell.font = fnt(bold=True, size=10)
    cell.border = thin
    cell.alignment = Alignment(
        horizontal="left" if wide else "center",
        vertical="center",
        wrap_text=wide,
        indent=1 if wide else 0,
    )
    unlock(cell)
    for c in range(4, end_col + 1):
        ws.cell(row, c).border = thin
        ws.cell(row, c).fill = fill(YELLOW)
        unlock(ws.cell(row, c))
    ws.row_dimensions[row].height = 28 if wide else 20


def _write_input_quote_meta(ws, tip_after_meta=True, defaults=None):
    """Carrier / Remark / Valid until under Currency footer."""
    defaults = defaults or {}
    for r in (38, META_TIP_ROW):
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row == r:
                try:
                    ws.unmerge_cells(str(rng))
                except Exception:
                    pass

    _footer_meta_row(ws, 37, "Carrier", defaults.get("carrier", "KE"))
    _footer_meta_row(
        ws,
        38,
        "Remark",
        defaults.get(
            "remark",
            "KEEP COOL / Maintained at 2-8°C at airline terminal",
        ),
        wide=True,
    )
    _footer_meta_row(ws, 39, "Valid until", defaults.get("valid", ""))

    if tip_after_meta:
        tip = META_TIP_ROW
        ws.merge_cells(start_row=tip, start_column=2, end_row=tip, end_column=10)
        tip_cell = ws.cell(
            tip,
            2,
            "Currency=Master CUR(J) | HKD·이미HKD→Ex.Rate=1 | Carrier·Remark·Valid→견적서",
        )
        tip_cell.font = fnt(size=9, color=MUTED)
        tip_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[tip].height = 22
        lock(tip_cell)


def _write_quote_meta(ws):
    """Shipment meta row + Remark above Charges."""
    IN = "입력"
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row in (9, 23):
            try:
                ws.unmerge_cells(str(rng))
            except Exception:
                pass

    for col, text in ((2, "Carrier"), (4, "Valid until")):
        cell = ws.cell(9, col, text)
        cell.font = fnt(bold=True, size=10)
        cell.fill = fill(LIGHT)
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
        lock(cell)

    c_cell = ws.cell(9, 3, f"='{IN}'!{CARRIER_CELL}")
    c_cell.font = fnt(bold=True, size=10)
    c_cell.fill = fill(WHITE)
    c_cell.border = thin
    c_cell.alignment = Alignment(horizontal="center", vertical="center")
    lock(c_cell)

    ws.merge_cells("E9:G9")
    v_cell = ws.cell(9, 5, f"='{IN}'!{VALID_CELL}")
    v_cell.font = fnt(bold=True, size=10)
    v_cell.fill = fill(WHITE)
    v_cell.border = thin
    v_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lock(v_cell)
    for c in (6, 7):
        ws.cell(9, c).border = thin
        ws.cell(9, c).fill = fill(WHITE)
        lock(ws.cell(9, c))
    ws.row_dimensions[9].height = 20

    lab = ws.cell(23, 2, "Remark")
    lab.font = fnt(bold=True, size=10, color=WHITE)
    lab.fill = fill(SLATE)
    lab.border = thin
    lab.alignment = Alignment(horizontal="center", vertical="center")
    lock(lab)
    ws.merge_cells("C23:G23")
    rem = ws.cell(23, 3, f"='{IN}'!{REMARK_CELL}")
    rem.font = fnt(size=10)
    rem.fill = fill(LIGHT)
    rem.border = thin
    rem.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    lock(rem)
    for c in range(4, 8):
        ws.cell(23, c).fill = fill(LIGHT)
        ws.cell(23, c).border = thin
        lock(ws.cell(23, c))
    ws.row_dimensions[23].height = 28


def _charge_line(ws, row, item, unit, ref_formula, editable_item=False):
    """Write one cost line matching baseline 참고(I)/예외(J) layout."""
    # E:F item, G:H unit, I ref, J ovr
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row == row and rng.min_col in (5, 7):
            try:
                ws.unmerge_cells(str(rng))
            except Exception:
                pass

    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    cell = ws.cell(row, 5, item)
    cell.font = fnt(bold=True, size=11)
    cell.fill = fill(INPUT_BG if editable_item else WHITE)
    cell.border = thin
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    (unlock if editable_item else lock)(cell)
    ws.cell(row, 6).border = thin
    ws.cell(row, 6).fill = fill(INPUT_BG if editable_item else WHITE)

    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    cell = ws.cell(row, 7, unit)
    cell.font = fnt(size=10)
    cell.fill = fill(LIGHT)
    cell.border = thin
    cell.alignment = Alignment(horizontal="center", vertical="center")
    lock(cell)
    ws.cell(row, 8).border = thin
    ws.cell(row, 8).fill = fill(LIGHT)

    ref = ws.cell(row, 9, ref_formula)
    ref.fill = fill(WHITE)
    ref.font = fnt(bold=True, size=10)
    ref.alignment = Alignment(horizontal="right", vertical="center")
    ref.number_format = "#,##0.00"
    ref.border = thin
    lock(ref)

    ovr = ws.cell(row, 10, None)
    ovr.fill = fill(YELLOW)
    ovr.font = fnt(bold=True, size=10)
    ovr.alignment = Alignment(horizontal="right", vertical="center")
    ovr.number_format = "#,##0.00"
    ovr.border = thin
    unlock(ovr)


def patch_input(ws, routes):
    unprotect(ws)

    # FX next to Currency block (after we move Currency). Place now at free area near tip.
    # Put Ex.Rate at H3/I3 (outside main cargo header merges carefully — E3:M3 is merged tip)
    # Use B30/C30 area after rebuild of bottom.

    # Update Break / Air Rate / FSC / SSC for -45 column layout
    ws["C21"] = (
        '=IFERROR(IF(C20="","",IF(C20<Master_DB!$B$9,"-45",'
        'IF(C20>=Master_DB!$B$12,"+1000",IF(C20>=Master_DB!$B$11,"+500",'
        'IF(C20>=Master_DB!$B$10,"+100","+45"))))),"")'
    )
    ws["C22"] = (
        '=IFERROR(IF(OR(C5="",C20=""),"",'
        f'IF(C20<Master_DB!$B$9,VLOOKUP(C5,{AIR_RNG},3,FALSE),'
        f'IF(C20>=Master_DB!$B$12,VLOOKUP(C5,{AIR_RNG},7,FALSE),'
        f'IF(C20>=Master_DB!$B$11,VLOOKUP(C5,{AIR_RNG},6,FALSE),'
        f'IF(C20>=Master_DB!$B$10,VLOOKUP(C5,{AIR_RNG},5,FALSE),'
        f'VLOOKUP(C5,{AIR_RNG},4,FALSE)))))),"")'
    )
    ws["C23"] = f'=IFERROR(IF(C5="","",VLOOKUP(C5,{AIR_RNG},2,FALSE)),"")'

    # Expand existing local INDEX ranges A21:A23 -> A21:A40
    for addr in ("G21", "I21", "G22", "I22", "G23", "I23"):
        v = ws[addr].value
        if isinstance(v, str):
            ws[addr].value = (
                v.replace("Master_DB!$A$21:$A$23", LOCAL_A)
                .replace("Master_DB!$B$21:$B$23", LOCAL_B)
                .replace("Master_DB!$C$21:$C$23", LOCAL_C)
                .replace("Master_DB!$D$21:$D$23", LOCAL_D)
                .replace("Master_DB!$A$16:$H$50", AIR_RNG)
                .replace("Master_DB!$A$16:$I$50", AIR_RNG)
            )

    # Clear old TOTAL / Currency block (will rewrite lower)
    # Unmerge TOTAL block E25:H26 / I25:J26 if present
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= 24:
            try:
                ws.unmerge_cells(str(rng))
            except Exception:
                pass

    # Wipe rows 24-40 content in charge/total area (E-J and B-D footer)
    for r in range(24, 42):
        for c in range(2, 11):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = PatternFill()
            cell.border = Border()

    # Restore left-panel FSC/SSC rates (I19/I20 and 견적서 depend on C24/C25).
    # Must run AFTER wipe — earlier wipe was clearing these and blanking Quote FSC/SSC.
    for r, label, col_idx in (
        (24, "FSC /kg", 8),
        (25, "SSC /kg", 9),
    ):
        lab = ws.cell(r, 2, label)
        lab.font = fnt(bold=True, size=10)
        lab.fill = fill(LIGHT)
        lab.border = thin
        lab.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lock(lab)
        rate = ws.cell(
            r,
            3,
            f'=IFERROR(IF(C5="","",VLOOKUP(C5,{AIR_RNG},{col_idx},FALSE)),"")',
        )
        rate.font = fnt(bold=True, size=11)
        rate.fill = fill(WHITE)
        rate.border = thin
        rate.alignment = Alignment(horizontal="center")
        rate.number_format = "0.00"
        lock(rate)

    # Terminal Charge (Master-linked)
    term_ref = (
        f'=IFERROR(IF(C20="","",MAX(INDEX({LOCAL_C},MATCH("Terminal Charge",{LOCAL_A},0))*C20,'
        f'INDEX({LOCAL_D},MATCH("Terminal Charge",{LOCAL_A},0)))),"")'
    )
    _charge_line(
        ws,
        24,
        "Terminal Charge",
        f'=IFERROR(INDEX({LOCAL_B},MATCH("Terminal Charge",{LOCAL_A},0)),"")',
        term_ref,
    )

    # Other 1-6 — generic slots with helpful default names + optional auto refs
    others = [
        ("Other 1 (XRAY)", "Per KG", '=IFERROR(IF(C20="","",INDEX(Master_DB!$C$21:$C$40,MATCH("XRAY",Master_DB!$A$21:$A$40,0))*C20),"")'),
        ("Other 2 (CFS)", "Per KG", '=IFERROR(IF(C20="","",MAX(INDEX(Master_DB!$C$21:$C$40,MATCH("CFS",Master_DB!$A$21:$A$40,0))*C20,INDEX(Master_DB!$D$21:$D$40,MATCH("CFS",Master_DB!$A$21:$A$40,0)))),"")'),
        ("Other 3 (Pickup)", "Per Shipment", '=IFERROR(INDEX(Master_DB!$D$21:$D$40,MATCH("Pickup (temp)",Master_DB!$A$21:$A$40,0)),"")'),
        ("Other 4 (Export)", "Per Entry", '=IFERROR(INDEX(Master_DB!$D$21:$D$40,MATCH("Export declaration",Master_DB!$A$21:$A$40,0))*F5,"")'),
        ("Other 5 (RE-PACK)", "Per PLT", '=IFERROR(IF(SUM(C11:L11)=0,"",INDEX(Master_DB!$C$21:$C$40,MATCH("RE-PACKING",Master_DB!$A$21:$A$40,0))*SUM(C11:L11)),"")'),
        ("Other 6 (Gate/etc)", "Manual", None),
    ]
    for i, (name, unit, formula) in enumerate(others):
        _charge_line(ws, 25 + i, name, unit, formula, editable_item=True)

    # TOTAL APPX
    parts = [f'IF(J{r}="",I{r},J{r})' for r in range(CH_FIRST, CH_LAST + 1)]
    # Currency=HKD -> * Ex.Rate; else keep
    conv = f'IF({CUR_CELL}="HKD",{FX_CELL},1)'
    tot_formula = f'=IFERROR(({"+".join(parts)})*{conv},"")'

    ws.merge_cells(f"E{TOT_ROW}:H{TOT_ROW + 1}")
    cell = ws.cell(TOT_ROW, 5, "TOTAL APPX. AMOUNT")
    cell.fill = fill(ORANGE)
    cell.font = fnt(bold=True, size=11, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin
    lock(cell)
    for r in range(TOT_ROW, TOT_ROW + 2):
        for c in range(5, 9):
            ws.cell(r, c).fill = fill(ORANGE)
            ws.cell(r, c).border = thin
            lock(ws.cell(r, c))

    ws.merge_cells(f"I{TOT_ROW}:J{TOT_ROW + 1}")
    tot = ws[TOT_AMT]
    tot.value = tot_formula
    tot.fill = fill(INPUT_BG)
    tot.font = fnt(bold=True, size=18)
    tot.alignment = Alignment(horizontal="center", vertical="center")
    tot.number_format = "#,##0.00"
    tot.border = thin
    lock(tot)
    for r in range(TOT_ROW, TOT_ROW + 2):
        for c in (9, 10):
            ws.cell(r, c).fill = fill(INPUT_BG)
            ws.cell(r, c).border = thin
            lock(ws.cell(r, c))

    # Footer: Currency (auto from Master CUR) / Ex.Rate / Route / C.W.
    foot = TOT_ROW + 2  # 33
    # Currency follows Route → Master!CUR (J). Unlock so desk can override if needed.
    cur_formula = f'=IFERROR(IF(C5="","",VLOOKUP(C5,{AIR_RNG},{CUR_COL},FALSE)),"USD")'
    labels = [
        (foot, "Currency", cur_formula, True),
        (foot + 1, "Ex.Rate", 7.8, True),
        (foot + 2, "Route", '=IFERROR(C5,"")', False),
        (foot + 3, "C.W.", '=IFERROR(C20,"")', False),
    ]
    for r, label, val, is_input in labels:
        lab = ws.cell(r, 2, label)
        lab.font = fnt(bold=True, size=10)
        lab.border = thin
        lab.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lock(lab)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        cell = ws.cell(r, 3, val)
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if is_input:
            if label == "Currency":
                cell.fill = fill(GREEN_BG)
                cell.font = fnt(bold=True, size=11, color=GREEN)
                unlock(cell)  # override OK
            else:
                cell.fill = fill(YELLOW)
                cell.font = fnt(bold=True, size=11)
                unlock(cell)
                cell.number_format = "0.0"
        else:
            cell.fill = fill(GREEN_BG if label == "C.W." else LIGHT)
            cell.font = fnt(
                bold=True if label == "C.W." else False,
                size=11,
                color=GREEN if label == "C.W." else NAVY,
            )
            lock(cell)
            if label == "C.W.":
                cell.number_format = "0.00"

    _write_input_quote_meta(ws, tip_after_meta=True)

    # Route + Currency dropdowns
    ws.data_validations.dataValidation = []
    if routes:
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(routes) + '"',
            allow_blank=False,
            showErrorMessage=False,
        )
        dv.add("C5")
        ws.add_data_validation(dv)
    dv_cur = DataValidation(
        type="list",
        formula1='"USD,HKD"',
        allow_blank=False,
        showErrorMessage=False,
    )
    dv_cur.add(CUR_CELL)
    ws.add_data_validation(dv_cur)

    protect(ws)


def patch_quote(ws):
    unprotect(ws)
    IN = "입력"

    # Currency ref -> C33, TOTAL -> I31, expand charge lines
    # Clear old charges area from row 24 down and rebuild
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= 24:
            try:
                ws.unmerge_cells(str(rng))
            except Exception:
                pass

    for r in range(24, 55):
        for c in range(2, 8):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = PatternFill()
            cell.border = Border()

    # Shipment currency already G7 = 입력!C27 — update to C33
    ws["G7"] = f"='{IN}'!{CUR_CELL}"
    _write_quote_meta(ws)

    ws.merge_cells("B24:G24")
    cell = ws.cell(24, 2, "Charges")
    cell.fill = fill(NAVY)
    cell.font = fnt(bold=True, size=11, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lock(cell)
    for c in range(2, 8):
        ws.cell(24, c).fill = fill(NAVY)
        ws.cell(24, c).border = thin
        lock(ws.cell(24, c))

    ws.merge_cells("B25:E25")
    cell = ws.cell(25, 2, "Description")
    cell.fill = fill(SLATE)
    cell.font = fnt(bold=True, size=10, color=WHITE)
    cell.alignment = Alignment(horizontal="center")
    lock(cell)
    for c in range(2, 6):
        ws.cell(25, c).fill = fill(SLATE)
        ws.cell(25, c).border = thin
        lock(ws.cell(25, c))

    ws.merge_cells("F25:G25")
    cell = ws.cell(25, 6, "Amount")
    cell.fill = fill(SLATE)
    cell.font = fnt(bold=True, size=10, color=WHITE)
    cell.alignment = Alignment(horizontal="center")
    lock(cell)
    ws.cell(25, 7).fill = fill(SLATE)
    ws.cell(25, 7).border = thin
    lock(ws.cell(25, 7))

    qr = 26
    for src in range(CH_FIRST, CH_LAST + 1):
        ws.merge_cells(start_row=qr, start_column=2, end_row=qr, end_column=5)
        name = ws.cell(qr, 2, f"='{IN}'!E{src}")
        name.font = fnt(bold=True, size=11)
        name.border = thin
        name.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        name.fill = fill(LIGHT)
        lock(name)
        for c in (3, 4, 5):
            ws.cell(qr, c).border = thin
            lock(ws.cell(qr, c))

        ws.merge_cells(start_row=qr, start_column=6, end_row=qr, end_column=7)
        # applied amount already FX-converted on 입력 TOTAL; line amounts show raw then
        # apply same conversion for display consistency
        amt = (
            f"=IFERROR(IF('{IN}'!J{src}=\"\",\"\",'{IN}'!J{src}),"
            f"IFERROR('{IN}'!I{src},\"\"))*IF('{IN}'!{CUR_CELL}=\"HKD\",'{IN}'!{FX_CELL},1)"
        )
        # Fix: IFERROR with two args wrongly. Use IF blank exception then ref.
        amt = (
            f"=IFERROR("
            f"IF('{IN}'!J{src}=\"\",'{IN}'!I{src},'{IN}'!J{src})"
            f"*IF('{IN}'!{CUR_CELL}=\"HKD\",'{IN}'!{FX_CELL},1),\"\")"
        )
        cell = ws.cell(qr, 6, amt)
        cell.fill = fill(LIGHT)
        cell.font = fnt(size=11)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = "#,##0.00"
        cell.border = thin
        lock(cell)
        ws.cell(qr, 7).fill = fill(LIGHT)
        ws.cell(qr, 7).border = thin
        ws.cell(qr, 7).alignment = Alignment(horizontal="right", vertical="center")
        lock(ws.cell(qr, 7))
        qr += 1

    ws.merge_cells(start_row=qr, start_column=2, end_row=qr, end_column=5)
    cell = ws.cell(qr, 2, "TOTAL APPX. AMOUNT")
    cell.fill = fill(ORANGE)
    cell.font = fnt(bold=True, size=11, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    lock(cell)
    for c in range(2, 6):
        ws.cell(qr, c).fill = fill(ORANGE)
        ws.cell(qr, c).border = thin
        lock(ws.cell(qr, c))

    ws.merge_cells(start_row=qr, start_column=6, end_row=qr + 1, end_column=7)
    cell = ws.cell(qr, 6, f"='{IN}'!{TOT_AMT}")
    cell.fill = fill(INPUT_BG)
    cell.font = fnt(bold=True, size=18)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = "#,##0.00"
    cell.border = thin
    lock(cell)
    for r in (qr, qr + 1):
        for c in (6, 7):
            ws.cell(r, c).fill = fill(INPUT_BG)
            ws.cell(r, c).border = thin
            lock(ws.cell(r, c))

    note_r = qr + 3
    ws.merge_cells(start_row=note_r, start_column=2, end_row=note_r, end_column=7)
    ws.cell(note_r, 2, "This quotation is indicative and subject to confirmation.")
    ws.cell(note_r, 2).font = fnt(size=9, color=MUTED)
    lock(ws.cell(note_r, 2))

    # PDF/print: fit one A4 page so charges are not clipped
    ws.print_area = f"B2:G{note_r}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True
    from openpyxl.worksheet.page import PageMargins

    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2
    )
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12

    protect(ws)


def build_guide(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 100
    ws["B2"] = "Guide — 범용 가견적 (CM UI 유지)"
    ws["B2"].font = fnt(bold=True, size=16)
    lines = [
        "",
        "1. 본질",
        "· Master_DB = 거의 고정 요율 (노란칸만 갱신)",
        "· 입력 = 화물 + 자동 C.W./Break + 참고(Master) / 예외(건별 수기) / Other(특수항목)",
        "· 견적서 = 입력 100% 참조 (화주 전달용)",
        "· 특수건(KEEP COOL 등)은 템플릿을 바꾸지 말고 예외·Other로 해결",
        "",
        "2. 과금 기준",
        "· Trucking = Rate × CBM (MIN) — kg 미적용",
        "· Terminal / CFS / XRAY = Rate × C.W.(kg) (MIN)",
        "· Handling / Doc / Pickup = 건별·BL·shipment",
        "· -45 = C.W. < 45kg / +45 이상 = 해당 weight break",
        "",
        "3. 통화",
        "· Master J열 CUR = 구간 요율 통화 (ICN*=USD, HKG*=HKD 기본)",
        "· 입력 Currency = Route VLOOKUP(CUR) 자동 (필요시 덮어쓰기)",
        "· Currency=USD → TOTAL 그대로 / HKD → Ex.Rate 곱함 (이미 HKD면 Ex.Rate=1)",
        "· 입력 Carrier / Remark / Valid until → 견적서에 표시 (KEEP COOL 등)",
        "",
        "4. 예시 — HKG-ICN KEEP COOL Chocolate (테스트 입력)",
        "· Route=HKG-ICN, B/L=1, #1=110×110×109 /1 /194.5 → C.W.≈220 Break=+100",
        "· Currency=HKD, Ex.Rate=1 (케이스 금액이 이미 HKD)",
        "· Handling 예외 321 / Doc 예외 15 / Trucking 예외 0 (해당 없으면)",
        "· Other 참고값 확인 후 필요시 예외 덮어쓰기 → 견적서 PDF",
        "",
        "5. ICN-HKG 6-pallet 샘플 (기본 데이터) TOTAL ≈ $3,728",
    ]
    for i, t in enumerate(lines):
        cell = ws.cell(3 + i, 2, t)
        cell.font = (
            fnt(bold=True, size=11, color=ORANGE)
            if t.startswith(("1.", "2.", "3.", "4.", "5."))
            else fnt(size=11)
        )


def verify():
    # CASE A: ICN-HKG classic sample dims from original verify engine
    air = {"MIN": 50, "u45": 5.5, 45: 4.5, 100: 3.8, 500: 3.2, 1000: 2.8, "FSC": 0.6, "SSC": 0.15}
    L, W, H, qty, gross = 110, 110, 150, 3, 400
    cbm = L * W * H * qty / 1_000_000
    cw = max(gross, cbm * 167)
    rate = air[500] if cw >= 500 else air[100]
    total_a = max(rate * cw, air["MIN"]) + air["FSC"] * cw + air["SSC"] * cw + 30 + 25 + max(15 * cbm, 80)
    assert abs(cbm - 5.445) < 0.001
    assert abs(cw - 909.315) < 0.01
    assert abs(total_a - 3728.47) < 0.02, total_a
    print(f"CASE A OK: APPX={total_a:.2f}")

    # Chocolate CW check (generic math, not template lock)
    cbm_c = 110 * 110 * 109 * 1 / 1_000_000
    cw_c = max(194.5, cbm_c * 167)
    assert abs(cw_c - 220.2563) < 0.01, cw_c
    print(f"Chocolate CW check OK: CBM={cbm_c:.4f} CW={cw_c:.4f}")
    return total_a


def main():
    verify()
    import sys

    # --currency-only / --meta-only: patch existing OUT (keeps chocolate fill)
    if "--currency-only" in sys.argv or "--meta-only" in sys.argv:
        src = OUT if OUT.exists() else OUT_ALT
        if not src.exists():
            raise SystemExit(f"No workbook to patch: {OUT}")
        wb = load_workbook(src)
        master = wb["Master_DB"]
        inp = wb[wb.sheetnames[1]]
        quote = wb[wb.sheetnames[2]]
        unprotect(master)
        unprotect(inp)
        unprotect(quote)
        _ensure_route_currency(master)
        master.column_dimensions["J"].width = 8
        for r in range(30, 36):
            v = master.cell(r, 1).value
            if isinstance(v, str) and ("ROUTE" in v or "구간" in v or "CUR" in v):
                master.cell(
                    r,
                    1,
                    "A16~ ROUTE | J열 CUR=USD/HKD | -45 / +45… | Trucking=CBM · Terminal/CFS/XRAY=C.W.",
                )
                master.cell(r, 1).alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
                master.row_dimensions[r].height = 22
                break
        cur_formula = f'=IFERROR(IF(C5="","",VLOOKUP(C5,{AIR_RNG},{CUR_COL},FALSE)),"USD")'
        cell = inp[CUR_CELL]
        cell.value = cur_formula
        cell.fill = fill(GREEN_BG)
        cell.font = fnt(bold=True, size=11, color=GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        unlock(cell)
        for addr in ("C22", "C23", "C24", "C25"):
            v = inp[addr].value
            if isinstance(v, str):
                inp[addr].value = v.replace("Master_DB!$A$16:$I$50", AIR_RNG)
        has_cur_dv = False
        for dv in inp.data_validations.dataValidation:
            if CUR_CELL in str(dv.sqref):
                has_cur_dv = True
                break
        if not has_cur_dv:
            dv_cur = DataValidation(
                type="list", formula1='"USD,HKD"', allow_blank=False, showErrorMessage=False
            )
            dv_cur.add(CUR_CELL)
            inp.add_data_validation(dv_cur)

        # Quote meta: Carrier / Remark / Valid until (chocolate defaults if empty)
        _write_input_quote_meta(
            inp,
            tip_after_meta=True,
            defaults={
                "carrier": "KE",
                "remark": "KEEP COOL / Maintained at 2-8°C at airline terminal",
                "valid": "31 Aug 2026",
            },
        )
        _write_quote_meta(quote)
        quote["G7"] = f"='입력'!{CUR_CELL}"

        if "Guide" in wb.sheetnames:
            build_guide(wb["Guide"])
        protect(master)
        protect(inp)
        protect(quote)
        wb.save(src)
        print(f"Live patch OK (currency+meta, kept inputs): {src}")
        for dest in (
            ROOT.parents[0] / "public" / "excel" / "WAC_Air_Quotation_Simulator.xlsx",
            Path.home() / "Downloads" / "WAC_Air_Quotation_Simulator.xlsx",
        ):
            try:
                dest.parent.mkdir(parents=True, exist_ok=True)
                dest.write_bytes(src.read_bytes())
                print(f"Copied -> {dest}")
            except OSError as e:
                print(f"Skip {dest}: {e}")
        return

    if not BASELINE.exists():
        raise SystemExit(f"Baseline missing: {BASELINE}")

    wb = load_workbook(BASELINE)
    # sheet names: Master_DB, 입력, 견적서
    master = wb["Master_DB"]
    inp = wb[wb.sheetnames[1]]
    quote = wb[wb.sheetnames[2]]

    routes = patch_master(master)
    patch_input(inp, routes)
    patch_quote(quote)

    if "Guide" in wb.sheetnames:
        del wb["Guide"]
    guide = wb.create_sheet("Guide", 3)
    build_guide(guide)

    saved = None
    for path in (OUT, OUT_ALT):
        try:
            wb.save(path)
            saved = path
            print(f"Wrote {path}")
            break
        except PermissionError:
            continue
    if not saved:
        raise SystemExit("Close Excel and retry")

    for dest in (
        ROOT.parents[0] / "public" / "excel" / "WAC_Air_Quotation_Simulator.xlsx",
        Path.home() / "Downloads" / "WAC_Air_Quotation_Simulator.xlsx",
    ):
        try:
            dest.parent.mkdir(parents=True, exist_ok=True)
            dest.write_bytes(saved.read_bytes())
            print(f"Copied -> {dest}")
        except OSError as e:
            print(f"Skip {dest}: {e}")


if __name__ == "__main__":
    main()
